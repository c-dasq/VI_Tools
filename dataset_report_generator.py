import argparse
import cv2
import matplotlib.pyplot as plt
import os
import pandas as pd
import requests
import shutil
import warnings
import weasyprint

from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from jinja2 import Template
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage, ImageDraw

def setup_report_folder(report_folder):
    if os.path.exists(report_folder):
        shutil.rmtree(report_folder)
    os.makedirs(report_folder)

def download_logo(logo_url, report_folder):
    logo_response = requests.get(logo_url, stream=True)
    if logo_response.status_code == 200:
        with open(os.path.join(report_folder, 'logo.jpg'), 'wb') as file:
            shutil.copyfileobj(logo_response.raw, file)

def round_corners(image, radius):
    circle = PILImage.new('L', (radius * 2, radius * 2), 0)
    draw = ImageDraw.Draw(circle)
    draw.ellipse((0, 0, radius * 2, radius * 2), fill=255)
    alpha = PILImage.new('L', image.size, 255)
    w, h = image.size
    alpha.paste(circle.crop((0, 0, radius, radius)), (0, 0))
    alpha.paste(circle.crop((0, radius, radius, radius * 2)), (0, h - radius))
    alpha.paste(circle.crop((radius, 0, radius * 2, radius)), (w - radius, 0))
    alpha.paste(circle.crop((radius, radius, radius * 2, radius * 2)), (w - radius, h - radius))
    image.putalpha(alpha)
    return image

def process_logo(report_folder):
    logo_image = PILImage.open(os.path.join(report_folder, 'logo.jpg')).convert("RGBA")
    logo_image = logo_image.resize((128, 128), PILImage.LANCZOS)
    logo_image = round_corners(logo_image, 20)
    logo_path = os.path.join(report_folder, 'logo_rounded.png')
    logo_image.save(logo_path)
    os.remove(os.path.join(report_folder, 'logo.jpg'))
    return logo_path

def fetch_data(url, api_key, datasetid, endpoint):
    headers = {"x-auth-token": api_key}
    response = requests.get(f"{url}/datasets/{datasetid}/{endpoint}", headers=headers, verify=False)
    return response.json()

def filter_files(files_response, fecha_inicio, fecha_fin, cantidad):
    def file_filter(file_info):
        if fecha_inicio and fecha_fin:
            file_created_at = file_info['created_at']
            return fecha_inicio <= file_created_at <= fecha_fin
        return True

    filtered_files = list(filter(file_filter, files_response))
    if cantidad:
        filtered_files = filtered_files[:cantidad]
    return filtered_files

def convert_timestamp(timestamp):
    return datetime.fromtimestamp(timestamp / 1000).strftime('%d-%b-%Y %H:%M')

def process_file(file_info, url, datasetid, report_folder, headers):
    file_id = file_info['_id']
    created_at = convert_timestamp(file_info['created_at'])
    file_path = os.path.join(report_folder, f"{file_id}.jpg")

    if not os.path.exists(file_path):
        file_content = requests.get(f"{url}/datasets/{datasetid}/files/{file_id}/download", headers=headers, verify=False).content
        with open(file_path, 'wb') as f:
            f.write(file_content)

    return {
        "file_id": file_id,
        "file_path": file_path,
        "created_at": created_at
    }

def gather_file_data(filtered_files, url, datasetid, report_folder, headers, labels_response):
    files_data = []
    existing_labels_data = []

    def process_and_collect(file_info):
        file_data = process_file(file_info, url, datasetid, report_folder, headers)
        files_data.append(file_data)

        file_id = file_data['file_id']
        existing_labels = [label['name'] for label in labels_response if label['file_id'] == file_id]
        label_counts = dict(Counter(existing_labels))
        label_counts['file_id'] = file_id
        label_counts['created_at'] = file_data['created_at']
        existing_labels_data.append(label_counts)

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(process_and_collect, file_info) for file_info in filtered_files]
        for future in as_completed(futures):
            future.result()

    return files_data, existing_labels_data

def create_excel_report(existing_labels_data, report_folder):
    existing_labels_df = pd.DataFrame(existing_labels_data).fillna(0)
    all_labels = list(set(existing_labels_df.columns) - {'file_id', 'created_at'})
    existing_labels_df = existing_labels_df.reindex(columns=['file_id', 'created_at'] + all_labels).fillna(0)

    excel_report_path = os.path.join(report_folder, 'dataset_report.xlsx')
    with pd.ExcelWriter(excel_report_path, engine='openpyxl') as writer:
        existing_labels_df.to_excel(writer, sheet_name='Existing Labels', index=False)

    workbook = load_workbook(excel_report_path)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        worksheet.sheet_view.showGridLines = False

    return excel_report_path, workbook, existing_labels_df, all_labels

def generate_pie_chart(label_counts, report_folder):
    total = sum(label_counts.values())
    plt.figure(figsize=(10, 7))

    def autopct_format(values):
        def my_format(pct):
            absolute = int(round(pct * total / 100.0))
            return f'{pct:.1f}%\n({absolute})'
        return my_format

    wedges, texts, autotexts = plt.pie(
        label_counts.values(),
        labels=label_counts.keys(),
        autopct=autopct_format(list(label_counts.values())),
        startangle=140
    )

    for text in texts:
        text.set_fontsize(12)
    for autotext in autotexts:
        autotext.set_fontsize(12)

    plt.axis('equal')
    plt.title('Percentage of Each Tag Detected', fontsize=16)
    pie_chart_path = os.path.join(report_folder, 'pie_chart.png')
    plt.savefig(pie_chart_path)
    plt.close()
    return pie_chart_path

def add_pie_chart_to_excel(workbook, pie_chart_path, excel_report_path):
    pie_chart_worksheet = workbook.create_sheet('Pie Chart')
    img = OpenpyxlImage(pie_chart_path)
    pie_chart_worksheet.add_image(img, 'A1')
    workbook.save(excel_report_path)

def process_image(file_info, labels_response, report_folder):
    file_id = file_info['_id']
    file_path = os.path.join(report_folder, f"{file_id}.jpg")
    labels_for_file = [label for label in labels_response if label['file_id'] == file_id]
    image = cv2.imread(file_path)

    for label in labels_for_file:
        bbox = label['bndbox']
        cv2.rectangle(image, (bbox['xmin'], bbox['ymin']), (bbox['xmax'], bbox['ymax']), (0, 255, 0), 2)
        cv2.putText(image, label['name'], (bbox['xmin'], bbox['ymin'] - 10), cv2.FONT_HERSHEY_SIMPLEX, 2.0, (0, 255, 0), 3)

    height, width = image.shape[:2]
    max_dimension = 800
    if max(height, width) > max_dimension:
        scaling_factor = max_dimension / float(max(height, width))
        new_size = (int(width * scaling_factor), int(height * scaling_factor))
        image = cv2.resize(image, new_size, interpolation=cv2.INTER_AREA)

    output_image_path = os.path.join(report_folder, f"{file_id}_detected.jpg")
    cv2.imwrite(output_image_path, image)

def process_detection_images(filtered_files, labels_response, report_folder):
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(process_image, file_info, labels_response, report_folder) for file_info in filtered_files]
        for future in as_completed(futures):
            future.result()

def generate_html_report(html_template, context, report_folder):
    template = Template(html_template)
    html_report = template.render(context)
    html_report_path = os.path.join(report_folder, 'dataset_report.html')
    with open(html_report_path, 'w') as f:
        f.write(html_report)
    return html_report_path

def generate_pdf_report(html_report_path, report_folder):
    pdf_report_path = os.path.join(report_folder, 'dataset_report.pdf')
    weasyprint.HTML(html_report_path).write_pdf(pdf_report_path)
    return pdf_report_path

def calculate_label_statistics(filtered_files, labels_response):
    label_image_counts = Counter()
    total_files = len(filtered_files)
    
    for file_info in filtered_files:
        file_id = file_info['_id']
        labels_for_file = {label['name'] for label in labels_response if label['file_id'] == file_id}
        for label in labels_for_file:
            label_image_counts[label] += 1
    
    return label_image_counts, total_files

def generate_dataset_report(url, api_key, datasetid, cantidad=None, fecha_inicio=None, fecha_fin=None):
    
    warnings.filterwarnings("ignore")
    report_folder = 'dataset_report'
    logo_url = 'https://is1-ssl.mzstatic.com/image/thumb/Purple211/v4/e4/8b/ff/e48bff38-6cd1-6646-624d-796fd8c80f44/AppIconLI-0-0-1x_U007emarketing-0-7-0-85-220.png/512x512bb.jpg'

    setup_report_folder(report_folder)
    download_logo(logo_url, report_folder)
    logo_path = process_logo(report_folder)

    headers = {"x-auth-token": api_key}
    tags_response = fetch_data(url, api_key, datasetid, 'tags')
    labels_response = fetch_data(url, api_key, datasetid, 'object-labels')
    files_response = fetch_data(url, api_key, datasetid, 'files')

    fecha_inicio_dt = fecha_fin_dt = None
    if fecha_inicio:
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%d-%b-%Y')
        fecha_inicio = fecha_inicio_dt.timestamp() * 1000
    if fecha_fin:
        fecha_fin_dt = datetime.strptime(fecha_fin, '%d-%b-%Y')
        fecha_fin = fecha_fin_dt.timestamp() * 1000

    filtered_files = filter_files(files_response, fecha_inicio, fecha_fin, cantidad)
    files_data, existing_labels_data = gather_file_data(filtered_files, url, datasetid, report_folder, headers, labels_response)

    label_image_counts, total_files = calculate_label_statistics(filtered_files, labels_response)

    excel_report_path, workbook, existing_labels_df, all_labels = create_excel_report(existing_labels_data, report_folder)

    label_counts = existing_labels_df[all_labels].sum().to_dict()
    pie_chart_path = generate_pie_chart(label_counts, report_folder)
    add_pie_chart_to_excel(workbook, pie_chart_path, excel_report_path)

    process_detection_images(filtered_files, labels_response, report_folder)

    html_template = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Dataset Report</title>
        <style>
            table {
                width: 100%;
                border-collapse: collapse;
                font-size: 10px;
            }
            table, th, td {
                border: 1px solid black;
            }
            th, td {
                padding: 8px;
                text-align: left;
            }
            th {
                background-color: #f2f2f2;
            }
            img {
                max-width: 100%;
                height: auto;
            }
            a {
                color: #0000EE;
                text-decoration: none;
            }
            a:hover {
                text-decoration: underline;
            }
            .logo {
                width: 64px;
                height: 64px;
                float: right;
                margin: 10px;
            }
            .report-date-range {
                font-size: 0.75em;
            }
            .explanatory-note {
                font-size: 0.8em;
                color: grey;
            }
        </style>
    </head>
    <body>
        <div class="logo">
            <img src="{{ logo_path }}" alt="Logo">
        </div>
        <h1>Dataset Report</h1>
        {% if report_date_range %}
        <h2 class="report-date-range">Report: {{ report_date_range }}</h2>
        {% endif %}
        <h2>Table of Contents</h2>
        <ul>
            <li><a href="#existing-labels">Existing Labels</a></li>
            <li><a href="#pie-chart">Pie Chart</a></li>
            <li><a href="#statistics">Statistics</a></li>
            <li><a href="#detection-images">Detection Images</a></li>
        </ul>

        <h2 id="existing-labels">Existing Labels</h2>
        {% if existing_labels_data %}
        <table>
            <thead>
                <tr>
                    <th>File ID</th>
                    <th>Created At</th>
                    {% for column in existing_labels_data[0].keys() if column not in ["file_id", "created_at"] %}
                    <th>{{ column }}</th>
                    {% endfor %}
                </tr>
            </thead>
            <tbody>
                {% for row in existing_labels_data %}
                <tr>
                    <td><a href="#{{ row.file_id }}">{{ row.file_id }}</a></td>
                    <td>{{ row.created_at }}</td>
                    {% for column in row.keys() if column not in ["file_id", "created_at"] %}
                    <td>{{ row[column] }}</td>
                    {% endfor %}
                </tr>
            {% endfor %}
        </tbody>
        </table>
        {% else %}
        <p>No existing labels data available.</p>
        {% endif %}

        <h2 id="pie-chart">Pie Chart</h2>
        <img src="{{ pie_chart_path }}" alt="Pie Chart">

        <h2 id="statistics">Statistics</h2>
        <table>
            <thead>
                <tr>
                    <th>Label</th>
                    <th>Count</th>
                    <th>Percentage</th>
                    <th>Total Images</th>
                </tr>
            </thead>
            <tbody>
                {% for label, count in label_counts.items() %}
                <tr>
                    <td>{{ label }}</td>
                    <td>{{ count }}</td>
                    <td>{{ (count / total_count * 100) | round(1) }}%</td>
                    <td>{{ label_image_counts[label] }}</td>
                </tr>
                {% endfor %}
            </tbody>
            <tfoot>
                <tr>
                    <th>Total</th>
                    <th>{{ total_count }}</th>
                    <th>100%</th>
                    <th>{{ total_files }}</th>
                </tr>
            </tfoot>
        </table>
        <p class="explanatory-note">
            * The grand total of the images may be greater than the partial sum of the totals, this is because many images may have multiple detections.
        </p>

        <h2 id="detection-images">Detection Images</h2>
        {% if detection_images %}
        {% for image in detection_images %}
        <div id="{{ image.file_id }}">
            <h3>File ID: {{ image.file_id }}</h3>
            <p>Created At: {{ image.created_at }}</p>
            <img src="{{ image.file_path }}" alt="Detection Image">
        </div>
        {% endfor %}
        {% else %}
        <p>No detection images available.</p>
        {% endif %}
    </body>
    </html>
    """

    report_date_range = None
    if fecha_inicio_dt and fecha_fin_dt:
        report_date_range = f"{fecha_inicio_dt.strftime('%d-%b-%Y')} - {fecha_fin_dt.strftime('%d-%b-%Y')}"

    context = {
        "existing_labels_data": existing_labels_df.to_dict(orient='records'),
        "pie_chart_path": os.path.abspath(pie_chart_path),
        "detection_images": [
            {
                "file_id": file_info['_id'],
                "file_path": os.path.abspath(os.path.join(report_folder, f"{file_info['_id']}_detected.jpg")),
                "created_at": convert_timestamp(file_info['created_at'])
            } for file_info in filtered_files
        ],
        "logo_path": os.path.abspath(logo_path),
        "label_counts": label_counts,
        "total_count": sum(label_counts.values()),
        "total_files": total_files,
        "label_image_counts": label_image_counts
    }

    if report_date_range:
        context["report_date_range"] = report_date_range

    html_report_path = generate_html_report(html_template, context, report_folder)
    pdf_report_path = generate_pdf_report(html_report_path, report_folder)

    print("Reporte HTML generado en dataset_report.html")
    print("Reporte PDF generado en dataset_report.pdf")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate the report")
    parser.add_argument("url", type=str, help="API URL")
    parser.add_argument("api_key", type=str, help="API key")
    parser.add_argument("datasetid", type=str, help="Dataset ID")
    parser.add_argument("--files", type=int, help="Number of files to process")
    parser.add_argument("--start-date", type=str, help="Start date (dd-MMM-yyyy)")
    parser.add_argument("--end-date", type=str, help="End date (dd-MMM-yyyy)")

    args = parser.parse_args()

    generate_dataset_report(
        args.url,
        args.api_key,
        args.datasetid,
        cantidad=args.files,
        fecha_inicio=args.start_date,
        fecha_fin=args.end_date
    )
